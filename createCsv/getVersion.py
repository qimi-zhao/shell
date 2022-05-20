# -*- coding: utf-8 -*-

import os
from openpyxl import load_workbook
from openpyxl import Workbook



# file global value
excel_home_path = "/Users/zhaolin/Documents/allFile/jita"
versionInfo={}

def getThisVersionInfo(path):
    global versionInfo
    wb = load_workbook(path)
    sheet = wb.get_sheet_by_name("変更一覧")

    col = 2
    value = ""
    while True:
        cellNmae = "B" + str(col)
        if sheet[cellNmae].value == None:
            break
        value = sheet[cellNmae].value
        col = col + 1
        
    versionInfo[path] = value


def getAllVersionInfo():
    global excel_home_path

    for root,dirs,files in os.walk(excel_home_path):
        for file in files:
            if file.endswith(".xlsx"):
                getThisVersionInfo(root + "/" + file)

def writeVersionInfo():
    global versionInfo
    wb = Workbook()
    sheet = wb.active
    sheet.title = "allVersionInfo"

    col = 1
    for key,value in versionInfo.items():
        cellKeyNmae = "A" + str(col)
        cellValueNmae = "B" + str(col)
        sheet[cellKeyNmae] = key
        sheet[cellValueNmae] = value

    wb.save('output.xlsx')

def main():
    # 读取所有版本信息
    getAllVersionInfo()
    # 输出结果
    writeVersionInfo()


if __name__ == "__main__":
    main()